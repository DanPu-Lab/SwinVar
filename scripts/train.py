from sched import scheduler
import pandas as pd
import numpy as np
import tables
import torch
import time
import os
import gc

from torch.utils.data import DataLoader
from torch_optimizer.lookahead import Lookahead
from openpyxl import load_workbook

from swinvar.postprocess.metrics_calculator import variant_df, calculate_metrics
from swinvar.preprocess.parameters import VARIANT_SIZE, GENOTYPE_SIZE
from swinvar.models.fine_tune import finetunning
from swinvar.preprocess.utils import setup_logger, check_directory
from swinvar.models.swin_var import SwinVar
from swinvar.models.focal_loss import MultiTaskLoss
from swinvar.models.dataset import CallingDataset


def create_alpha(labels, num_classes):
    class_counts = np.bincount(labels, minlength=num_classes)
    max_count = np.max(class_counts)
    class_weights = class_counts / max_count

    return torch.tensor(class_weights, dtype=torch.float32)


def train_model(args):

    start_all = time.time()

    output_path = os.path.join(args["output_path"], "train_moe", args["file"])
    check_directory(output_path)

    model_save_path = os.path.join(output_path, args["model_save_path"])

    if args["LoRA"]:
        output_path = os.path.join(args["output_path"], "train_moe", f"LoRA_{args["LoRA_file"]}_{args["ref_var_ratio"]}")
        check_directory(output_path)

    matplot_save_path = os.path.join(output_path, args["matplot_save_path"])
    hyperparams_outfile = os.path.join(output_path, args["hyperparams_log"])
    log_file = os.path.join(output_path, args["log_file_train"])
    checkpoint_path = os.path.join(output_path, "checkpoint.pth")
    logger = setup_logger(log_file) if not args["checkpoint"] else setup_logger(log_file, mode="a")

    # tables.set_blosc_max_threads(64)
    if isinstance(args["input_path"], list):
        train_inputs_files = []
        val_inputs_files = []
        for input_file in args["input_path"]:
            input_path = os.path.join(input_file, args["file"]) if not args["LoRA"] else os.path.join(input_file, args["LoRA_file"])
            all_files = os.listdir(input_path)
            train_inputs_files.extend(os.path.join(input_path, file) for file in all_files if "20." not in file)
            # train_inputs_files.extend(os.path.join(input_path, file) for file in all_files if "1." in file)
            val_inputs_files.extend(os.path.join(input_path, file) for file in all_files if "20." in file)
    else:
        input_path = os.path.join(args["input_path"], args["file"]) if not args["LoRA"] else os.path.join(input_file, args["LoRA_file"])
        all_files = os.listdir(input_path)
        train_inputs_files = [os.path.join(input_path, file) for file in all_files if "20." not in file]
        val_inputs_files = [os.path.join(input_path, file) for file in all_files if "20." in file]

    train_tables_data_list = [tables.open_file(file, "r") for file in train_inputs_files]
    val_tables_data_list = [tables.open_file(file, "r") for file in val_inputs_files]

    train_dataset = CallingDataset(train_tables_data_list)
    val_dataset = CallingDataset(val_tables_data_list)

    train_dataloader = DataLoader(
        train_dataset,
        batch_size=args["batch_size"] if not args["LoRA"] else args["lora_batch_size"],
        shuffle=True,
        pin_memory=True,
        num_workers=args["num_workers"],
    )

    val_dataloader = DataLoader(
        val_dataset,
        batch_size=args["batch_size"] if not args["LoRA"] else args["lora_batch_size"],
        shuffle=True,
        pin_memory=True,
        num_workers=args["num_workers"],
    )

    device = torch.device(f"cuda")
    model = SwinVar(
        feature_size=args["feature_size"],
        num_classes=args["num_classes"],
        embed_dim=args["embed_dim"],
        patch_size=args["patch_size"],
        window_size=args["window_size"],
        n_routed_experts=args["n_routed_experts"],
        n_activated_experts=args["n_activated_experts"],
        n_expert_groups=args["n_expert_groups"],
        n_limited_groups=args["n_limited_groups"],
        score_func=args["score_func"],
        route_scale=args["route_scale"],
        moe_inter_dim=args["moe_inter_dim"],
        n_shared_experts=args["n_shared_experts"],
        depths=args["depths"],
        num_heads=args["num_heads"],
        drop_rate=args["drop_rate"],
        drop_path_rate=args["drop_path_rate"],
        attn_drop_rate=args["attn_drop_rate"],
    ).to(device)

    if args["LoRA"]:
        model.load_state_dict(torch.load(model_save_path, map_location=device))
        # model = apply_lora_to_model(model, args["lora_r"], args["lora_alpha"])
        finetunning(model)
        model_save_path = os.path.join(output_path, f"LoRA_{args["model_save_path"]}")

    model.to(device)
    ft_params = sum(p.numel() for p in model.parameters() if p.requires_grad)
    total_params = sum(p.numel() for p in model.parameters())
    if args["LoRA"]:
        print(f"模型的总参数: {total_params}, 微调模型参数: {ft_params}")
    else:
        print(f"模型的总参数: {total_params}")
    # model.print_trainable_parameters()

    variant_1_labels_list = []
    variant_2_labels_list = []
    genotype_labels_list = []
    for tables_data in train_tables_data_list:
        variant_1_labels_list.append(tables_data.root.Variant_labels[:][:, 0])
        variant_2_labels_list.append(tables_data.root.Variant_labels[:][:, 1])
        genotype_labels_list.append(tables_data.root.Variant_labels[:][:, 2])

    all_variant_1_labels = np.concatenate(variant_1_labels_list, axis=0)
    all_variant_2_labels = np.concatenate(variant_2_labels_list, axis=0)
    all_genotype_labels = np.concatenate(genotype_labels_list, axis=0)

    variant_1_labels_counts = np.bincount(all_variant_1_labels, minlength=VARIANT_SIZE)
    variant_2_labels_counts = np.bincount(all_variant_2_labels, minlength=VARIANT_SIZE)
    genotype_labels_counts = np.bincount(all_genotype_labels, minlength=GENOTYPE_SIZE)
    
    if not args["checkpoint"]:
        logger.info(f"Train Labels_1 Counts: {variant_1_labels_counts}")
        logger.info(f"Train Labels_2 Counts: {variant_2_labels_counts}")
        logger.info(f"Train Genotype Labels Counts: {genotype_labels_counts}")

    alpha_variant_1 = create_alpha(all_variant_1_labels, VARIANT_SIZE).to(device)
    alpha_variant_2 = create_alpha(all_variant_2_labels, VARIANT_SIZE).to(device)
    alpha_genotype = create_alpha(all_genotype_labels, GENOTYPE_SIZE).to(device)

    alpha_variant_1 = torch.tensor([1, 1, 1, 1, 2, 2], dtype=torch.float32, device=device)
    alpha_variant_2 = torch.tensor([1, 1, 1, 1, 2, 2], dtype=torch.float32, device=device)
    alpha_genotype = torch.tensor([1, 2, 2, 4], dtype=torch.float32, device=device)
    gamma_variant_1 = torch.tensor([1, 1, 1, 1, 2, 2], dtype=torch.float32, device=device)
    gamma_variant_2 = torch.tensor([1, 1, 1, 1, 2, 2], dtype=torch.float32, device=device)
    gamma_genotype = torch.tensor([1, 2, 2, 4], dtype=torch.float32, device=device)
    class_loss = MultiTaskLoss(alphas=[alpha_variant_1, alpha_variant_2, alpha_genotype], gammas=[gamma_variant_1, gamma_variant_2, gamma_genotype], label_smoothing=0.1)

    base_optimizer = torch.optim.AdamW(
        model.parameters() if not args["LoRA"] else filter(lambda p: p.requires_grad, model.parameters()),
        lr=args["lr"] if not args["LoRA"] else args["lora_lr"],
        weight_decay=args["weight_decay"],
    )
    optimizer = Lookahead(base_optimizer, k=5, alpha=0.5)
    # scheduler = torch.optim.lr_scheduler.OneCycleLR(
    #     base_optimizer, max_lr=(args["lr"]*1),
    #     total_steps=len(train_dataloader) * args["epochs"],
    #     pct_start=args["pct_start"],
    #     anneal_strategy="cos",
    #     div_factor=args["factor"][0],
    #     final_div_factor=args["factor"][1],
    # )

    # 初始化
    best_val_loss = float("inf")
    best_val_acc = float("-inf")
    best_train_f1_score = float("-inf")
    best_train_f1_score_indel = float("-inf")
    best_val_f1_score = float("-inf")
    best_val_f1_score_snp = float("-inf")
    best_val_f1_score_indel = float("-inf")

    patience_counter = 0
    patience = args["patience"] if not args["LoRA"] else args["lora_patience"]
    patience_lr_scheduler = patience // 4
    train_loss_list, val_loss_list = [], []
    train_variant_1_acc_list, val_variant_1_acc_list = [], []
    train_variant_2_acc_list, val_variant_2_acc_list = [], []
    train_genotype_acc_list, val_genotype_acc_list = [], []

    train_sensitivity_list, val_sensitivity_list = [], []
    train_precision_list, val_precision_list = [], []
    train_f1_score_list, val_f1_score_list = [], []

    train_f1_score_snp_list, val_f1_score_snp_list = [], []
    train_f1_score_indel_list, val_f1_score_indel_list = [], []

    scaler = torch.amp.GradScaler("cuda")

    epochs = args["epochs"] if not args["LoRA"] else args["lora_epochs"]
    start_epoch = 0
    
    if args["checkpoint"]:
        checkpoint_dict = torch.load(checkpoint_path, weights_only=False, map_location=device)
        model.load_state_dict(checkpoint_dict["model_state_dict"])
        optimizer.optimizer.load_state_dict(checkpoint_dict["optimizer_state_dict"])

        for param_group, saved_group in zip(optimizer.optimizer.param_groups, checkpoint_dict["optimizer_state_dict"]["param_groups"]):
            param_group["lr"] = saved_group["lr"]
            param_group['weight_decay'] = saved_group['weight_decay']

        start_epoch = checkpoint_dict["epoch"]
        patience_counter = checkpoint_dict["patience_counter"]
        best_val_f1_score = checkpoint_dict["best_val_f1_score"]
        best_train_f1_score = checkpoint_dict["best_train_f1_score"]
        best_epoch = checkpoint_dict["best_epoch"]
        patience_lr_scheduler = checkpoint_dict["patience_lr_scheduler"]

        train_loss_list = checkpoint_dict["train_loss_list"]
        val_loss_list = checkpoint_dict["val_loss_list"]
        train_sensitivity_list = checkpoint_dict["train_sensitivity_list"]
        val_sensitivity_list = checkpoint_dict["val_sensitivity_list"]
        train_precision_list = checkpoint_dict["train_precision_list"]
        val_precision_list = checkpoint_dict["val_precision_list"]
        train_f1_score_list = checkpoint_dict["train_f1_score_list"]
        val_f1_score_list = checkpoint_dict["val_f1_score_list"]

    for epoch in range(start_epoch, epochs):

        start = time.time()

        model.train()

        train_loss, train_variant_1_acc, train_variant_2_acc, train_genotype_acc, train_num = 0.0, 0.0, 0.0, 0.0, 0
        (
            train_variant_1_predictions,
            train_variant_2_predictions,
            train_genotype_predictions,
            train_variant_1_labels,
            train_variant_2_labels,
            train_genotype_labels,
            train_chrom,
            train_pos,
            train_ref,
            train_indel_info,
        ) = ([] for _ in range(10))

        for batch in train_dataloader:
            features, variant_1_labels, variant_2_labels, genotype_labels, chrom, pos, ref, indel_info = batch

            features, variant_1_labels, variant_2_labels, genotype_labels = (
                features.to(device),
                variant_1_labels.reshape(-1).to(device),
                variant_2_labels.reshape(-1).to(device),
                genotype_labels.reshape(-1).to(device),
            )

            optimizer.zero_grad()

            with torch.amp.autocast("cuda"):
                variant_1_classification, variant_2_classification, genotype_classification = model(features)
                loss = class_loss([variant_1_classification, variant_2_classification, genotype_classification], [variant_1_labels, variant_2_labels, genotype_labels])

            scaler.scale(loss).backward()
            scaler.step(optimizer)
            scaler.update()

            train_variant_1_pred = torch.argmax(variant_1_classification, dim=-1)
            train_variant_2_pred = torch.argmax(variant_2_classification, dim=-1)
            train_genotype_pred = torch.argmax(genotype_classification, dim=-1)

            train_variant_1_predictions.extend(train_variant_1_pred.cpu().numpy())
            train_variant_2_predictions.extend(train_variant_2_pred.cpu().numpy())
            train_genotype_predictions.extend(train_genotype_pred.cpu().numpy())
            train_variant_1_labels.extend(variant_1_labels.cpu().numpy())
            train_variant_2_labels.extend(variant_2_labels.cpu().numpy())
            train_genotype_labels.extend(genotype_labels.cpu().numpy())
            train_chrom.extend(chrom)
            train_pos.extend(pos)
            train_ref.extend(ref)
            train_indel_info.extend(indel_info)

            train_loss += loss.item() * features.size(0)
            train_variant_1_acc += (train_variant_1_pred == variant_1_labels).sum().item()
            train_variant_2_acc += (train_variant_2_pred == variant_2_labels).sum().item()
            train_genotype_acc += (train_genotype_pred == genotype_labels).sum().item()
            train_num += features.size(0)

        val_loss, val_variant_1_acc, val_variant_2_acc, val_genotype_acc, val_num = 0.0, 0.0, 0.0, 0.0, 0
        (
            val_variant_1_predictions,
            val_variant_2_predictions,
            val_genotype_predictions,
            val_variant_1_labels,
            val_variant_2_labels,
            val_genotype_labels,
            val_chrom,
            val_pos,
            val_ref,
            val_indel_info,
        ) = ([] for _ in range(10))

        model.eval()
        with torch.no_grad():
            for batch in val_dataloader:
                features, variant_1_labels, variant_2_labels, genotype_labels, chrom, pos, ref, indel_info = batch
                features, variant_1_labels, variant_2_labels, genotype_labels = (
                    features.to(device),
                    variant_1_labels.reshape(-1).to(device),
                    variant_2_labels.reshape(-1).to(device),
                    genotype_labels.reshape(-1).to(device),
                )

                with torch.amp.autocast("cuda"):
                    variant_1_classification, variant_2_classification, genotype_classification = model(features)

                    loss = class_loss([variant_1_classification, variant_2_classification, genotype_classification], [variant_1_labels, variant_2_labels, genotype_labels])

                val_variant_1_pred = torch.argmax(variant_1_classification, dim=-1)
                val_variant_2_pred = torch.argmax(variant_2_classification, dim=-1)
                val_genotype_pred = torch.argmax(genotype_classification, dim=-1)

                val_variant_1_predictions.extend(val_variant_1_pred.cpu().numpy())
                val_variant_2_predictions.extend(val_variant_2_pred.cpu().numpy())
                val_genotype_predictions.extend(val_genotype_pred.cpu().numpy())
                val_variant_1_labels.extend(variant_1_labels.cpu().numpy())
                val_variant_2_labels.extend(variant_2_labels.cpu().numpy())
                val_genotype_labels.extend(genotype_labels.cpu().numpy())
                val_chrom.extend(chrom)
                val_pos.extend(pos)
                val_ref.extend(ref)
                val_indel_info.extend(indel_info)

                val_variant_1_acc += (val_variant_1_pred.detach() == variant_1_labels.detach()).sum().item()
                val_variant_2_acc += (val_variant_2_pred.detach() == variant_2_labels.detach()).sum().item()
                val_genotype_acc += (val_genotype_pred.detach() == genotype_labels.detach()).sum().item()
                val_loss += loss.item() * features.size(0)
                val_num += features.size(0)

        # loss
        train_loss_list.append(train_loss / train_num)
        train_variant_1_acc_list.append(train_variant_1_acc / train_num)
        train_variant_2_acc_list.append(train_variant_2_acc / train_num)
        train_genotype_acc_list.append(train_genotype_acc / train_num)

        val_loss_list.append(val_loss / val_num)
        val_variant_1_acc_list.append(val_variant_1_acc / val_num)
        val_variant_2_acc_list.append(val_variant_2_acc / val_num)
        val_genotype_acc_list.append(val_genotype_acc / val_num)

        train_df = variant_df(train_chrom, train_pos, train_ref, train_indel_info, train_variant_1_predictions, train_variant_1_labels, train_variant_2_predictions, train_variant_2_labels, train_genotype_predictions, train_genotype_labels)
        train_df_snp = train_df[~train_df["Label"].str.contains("Insert|Deletion", regex=True)]
        train_df_indel = train_df[train_df["Label"].str.contains("Insert|Deletion", regex=True)]
        train_df = train_df.drop_duplicates().reset_index(drop=True)

        val_df = variant_df(val_chrom, val_pos, val_ref, val_indel_info, val_variant_1_predictions, val_variant_1_labels, val_variant_2_predictions, val_variant_2_labels, val_genotype_predictions, val_genotype_labels)
        val_df_snp = val_df[~val_df["Label"].str.contains("Insert|Deletion", regex=True)]
        val_df_indel = val_df[val_df["Label"].str.contains("Insert|Deletion", regex=True)]
        val_df = val_df.drop_duplicates().reset_index(drop=True)

        # ALL
        train_variant_count = (train_df["Label_Genotype"] != 0).sum()
        val_variant_count = (val_df["Label_Genotype"] != 0).sum()

        train_variant, train_sensitivity, train_precision, train_f1_score = calculate_metrics(train_df)
        val_variant, val_sensitivity, val_precision, val_f1_score = calculate_metrics(val_df)

        train_sensitivity_list.append(train_sensitivity)
        train_precision_list.append(train_precision)
        train_f1_score_list.append(train_f1_score)

        val_sensitivity_list.append(val_sensitivity)
        val_precision_list.append(val_precision)
        val_f1_score_list.append(val_f1_score)

        # SNP
        train_variant_snp, train_sensitivity_snp, train_precision_snp, train_f1_score_snp = calculate_metrics(train_df_snp)
        val_variant_snp, val_sensitivity_snp, val_precision_snp, val_f1_score_snp = calculate_metrics(val_df_snp)

        train_f1_score_snp_list.append(train_f1_score_snp)
        val_f1_score_snp_list.append(val_f1_score_snp)

        # INDEL
        train_variant_indel, train_sensitivity_indel, train_precision_indel, train_f1_score_indel = calculate_metrics(train_df_indel)
        val_variant_indel, val_sensitivity_indel, val_precision_indel, val_f1_score_indel = calculate_metrics(val_df_indel)

        train_f1_score_indel_list.append(train_f1_score_indel)
        val_f1_score_indel_list.append(val_f1_score_indel)

        # 早停
        if (val_f1_score_list[-1] > best_val_f1_score) or (val_f1_score_list[-1] == best_val_f1_score and train_f1_score_list[-1] > best_train_f1_score):

            best_train_f1_score = train_f1_score_list[-1]

            best_val_loss = val_loss_list[-1]
            best_val_acc = val_variant_1_acc_list[-1]
            best_val_f1_score = val_f1_score_list[-1]
            best_epoch = epoch
            patience_counter = 0
            patience_lr_scheduler = patience // 4

            torch.save(
                model.state_dict(),
                model_save_path,
            )
            # train_df.to_csv(os.path.join(output_path, "train_df.csv"), index=False)
            # train_df_snp.to_csv(os.path.join(output_path, "train_df_snp.csv"), index=False)
            # train_df_indel.to_csv(os.path.join(output_path, "train_df_indel.csv"), index=False)
            # val_df.to_csv(os.path.join(output_path, "val_df.csv"), index=False)

        else:
            patience_counter += 1

        current_lr = optimizer.optimizer.param_groups[0]["lr"]

        logger.info(f"Epoch [{epoch+1}/{epochs}]: learning rate: {current_lr:g}")
        logger.info(
            f"Train Loss: {train_loss_list[-1]:.6f}\tTrain Genotype Acc: {train_genotype_acc_list[-1]:.6f}\tTrain Variant 1 Acc: {train_variant_1_acc_list[-1]:.6f}\tTrain Variant 2 Acc: {train_variant_2_acc_list[-1]:.6f}"
        )
        logger.info(f"Val   Loss: {val_loss_list[-1]:.6f}\tVal Genotype Acc: {val_genotype_acc_list[-1]:.6f}\tVal Variant 1 Acc: {val_variant_1_acc_list[-1]:.6f}\tVal Variant 2 Acc: {val_variant_2_acc_list[-1]:.6f}")
        logger.info(f"patience: {patience_counter}\n")
        logger.info(f"Train----Total number of variants: {train_variant_count}")
        logger.info(f"Val------Total number of variants: {val_variant_count}")
        logger.info(f"Train----True negative: {train_variant["True negative"]}\tTrue positive: {train_variant['True positive']}\tFalse positive: {train_variant['False positive']}\tFalse negative: {train_variant['False negative']}")
        logger.info(f"Val------True negative: {val_variant["True negative"]}\tTrue positive: {val_variant['True positive']}\tFalse positive: {val_variant['False positive']}\tFalse negative: {val_variant['False negative']}")
        logger.info(f"Train----Sensitivity: {train_sensitivity:.6f}\tPrecision: {train_precision:.6f}\tF1 Score: {train_f1_score:.6f}")
        logger.info(f"Val------Sensitivity: {val_sensitivity:.6f}\tPrecision: {val_precision:.6f}\tF1 Score: {val_f1_score:.6f}")
        logger.info(f"SNP:")
        logger.info(f"Train----True negative: {train_variant_snp["True negative"]}\tTrue positive: {train_variant_snp['True positive']}\tFalse positive: {train_variant_snp['False positive']}\tFalse negative: {train_variant_snp['False negative']}")
        logger.info(f"Val------True negative: {val_variant_snp["True negative"]}\tTrue positive: {val_variant_snp['True positive']}\tFalse positive: {val_variant_snp['False positive']}\tFalse negative: {val_variant_snp['False negative']}")
        logger.info(f"Train----Sensitivity:{train_sensitivity_snp:.6f}\tPrecision:{train_precision_snp:.6f}\tF1 Score:{train_f1_score_snp:.6f}")
        logger.info(f"Val------Sensitivity:{val_sensitivity_snp:.6f}\tPrecision:{val_precision_snp:.6f}\tF1 Score:{val_f1_score_snp:.6f}")
        logger.info(f"INDEL:")
        logger.info(f"Train----True negative: {train_variant_indel["True negative"]}\tTrue positive: {train_variant_indel['True positive']}\tFalse positive: {train_variant_indel['False positive']}\tFalse negative: {train_variant_indel['False negative']}")
        logger.info(f"Val------True negative: {val_variant_indel["True negative"]}\tTrue positive: {val_variant_indel['True positive']}\tFalse positive: {val_variant_indel['False positive']}\tFalse negative: {val_variant_indel['False negative']}")
        logger.info(f"Train----Sensitivity:{train_sensitivity_indel:.6f}\tPrecision:{train_precision_indel:.6f}\tF1 Score:{train_f1_score_indel:.6f}")
        logger.info(f"Val------Sensitivity:{val_sensitivity_indel:.6f}\tPrecision:{val_precision_indel:.6f}\tF1 Score:{val_f1_score_indel:.6f}")
        logger.info(f"time: {(time.time()-start)//60:.0f}m{(time.time()-start)%60:.0f}s")
        logger.info(f"{"-"*100}")

        if patience_counter > 0 and patience_counter % patience_lr_scheduler == 0:
            patience_lr_scheduler += patience_lr_scheduler
            
            for param_group in optimizer.optimizer.param_groups:
                param_group["lr"] *= 0.1
                
        if patience_counter >= patience:
            logger.info(f"\n{'*'*120}")
            logger.info(
                f"Early stopping at epoch: {epoch + 1}! Saving model at {best_epoch + 1} with Train F1 Score: {best_train_f1_score:.6f}\tVal F1 Score: {best_val_f1_score:.6f}"
            )
            logger.info(f"{'*'*120}")
            break

        if epoch + 1 == epochs:
            logger.info(f"\n{'*'*120}")
            logger.info(
                f"Saving model at epoch: {best_epoch + 1} with Train F1 Score: {best_train_f1_score:.6f}\tVal F1 Score: {best_val_f1_score:.6f}"
            )
            logger.info(f"{'*'*120}\n")

        # checkpoint
        checkpoint_dict = {
            "epoch":epoch+1,
            "model_state_dict": model.state_dict(),
            "optimizer_state_dict": optimizer.optimizer.state_dict(),
            "patience_counter": patience_counter,
            "best_val_f1_score": best_val_f1_score,
            "best_train_f1_score": best_train_f1_score,
            "best_epoch": best_epoch,
            "patience_lr_scheduler": patience_lr_scheduler,
            "train_loss_list": train_loss_list,
            "val_loss_list": val_loss_list,
            "train_sensitivity_list": train_sensitivity_list,
            "val_sensitivity_list": val_sensitivity_list,
            "train_precision_list": train_precision_list,
            "val_precision_list": val_precision_list,
            "train_f1_score_list": train_f1_score_list,
            "val_f1_score_list": val_f1_score_list,

        }
        torch.save(checkpoint_dict, checkpoint_path)

        torch.cuda.empty_cache()
        gc.collect()

    for tables_data in train_tables_data_list:
        tables_data.close()

    for tables_data in val_tables_data_list:
        tables_data.close()

    hyperparams = {
        "model": model.__class__.__name__,
        "optimizer": type(optimizer).__name__,
        "best_val_loss": best_val_loss,
        "best_val_acc": best_val_acc,
        "best_val_f1_score_all": best_val_f1_score,
        "best_val_f1_score_snp": best_val_f1_score_snp,
        "best_val_f1_score_indel": best_val_f1_score_indel,
        "train_ratio": args["train_ratio"],
        "seed": args["seed"],
        "epochs": args["epochs"],
        "best_epoch": epoch,
        "batch_size": args["batch_size"],
        "depths": args["depths"],
        "num_heads": args["num_heads"],
        "embed_dim": args["embed_dim"],
        "patch_size": args["patch_size"],
        "window_size": args["window_size"],
        "n_routed_experts": args["n_routed_experts"],
        "n_activated_experts": args["n_activated_experts"],
        "n_expert_groups": args["n_expert_groups"],
        "n_limited_groups": args["n_limited_groups"],
        "score_func": args["score_func"],
        "route_scale": args["route_scale"],
        "moe_inter_dim": args["moe_inter_dim"],
        "n_shared_experts": args["n_shared_experts"],
        "drop_rate": args["drop_rate"],
        "drop_path_rate": args["drop_path_rate"],
        "attn_drop_rate": args["attn_drop_rate"],
        "learning_rate": args["lr"],
        "weight_decay": args["weight_decay"],
        "pct_start": args["pct_start"],
    }

    hyperparams_df = pd.DataFrame([hyperparams])

    if os.path.exists(hyperparams_outfile):

        book = load_workbook(hyperparams_outfile)
        if "Transformer" in book.sheetnames:
            sheet = book["Transformer"]
            start_row = sheet.max_row
        else:
            start_row = 0
        with pd.ExcelWriter(
            hyperparams_outfile,
            mode="a",
            engine="openpyxl",
            if_sheet_exists="overlay",
        ) as writer:
            hyperparams_df.to_excel(writer, sheet_name="Transformer", index=False, header=not ("Transformer" in load_workbook(hyperparams_outfile).sheetnames), startrow=start_row)
    else:
        with pd.ExcelWriter(
            hyperparams_outfile,
            mode="w",
            engine="openpyxl",
        ) as writer:
            hyperparams_df.to_excel(writer, sheet_name="Transformer", index=False, header=True)

    del model, base_optimizer, optimizer, train_dataset, val_dataset, train_dataloader, val_dataloader
    gc.collect()
    torch.cuda.empty_cache()

    consum_all = time.time() - start_all
    hours, remainder = divmod(consum_all, 3600)
    minutes, seconds = divmod(remainder, 60)
    logger.info(f"train time: {hours:.0f}h{minutes:.0f}m{seconds:.0f}s")

    train_process = pd.DataFrame(
        data={
            "epoch": range(1, epoch + 2),
            "train_loss_list": train_loss_list,
            "val_loss_list": val_loss_list,
            "train_sensitivity_list": train_sensitivity_list,
            "val_sensitivity_list": val_sensitivity_list,
            "train_precision_list": train_precision_list,
            "val_precision_list": val_precision_list,
            "train_f1_score_list": train_f1_score_list,
            "val_f1_score_list": val_f1_score_list,
        },
    )
    train_process.to_excel(f"{output_path}/train_data.xlsx", index=False)
