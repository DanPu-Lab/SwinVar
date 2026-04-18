import os
import time
import pandas as pd
from openpyxl import load_workbook

from swinvar.preprocess.utils import setup_logger, check_directory


class TrainingConfig:
    """训练配置管理类"""
    
    def __init__(self, args):
        self.args = args
        self._setup_paths()
        self._setup_logging()
        self._validate_config()
    
    def _setup_paths(self):
        """设置路径配置"""
        base_output_path = os.path.join(self.args["output_path"], "train_moe", self.args["file"])
        
        if self.args["ft"]:
            self.output_path = os.path.join(
                self.args["output_path"], 
                "train_moe", 
                f"ft_{self.args['ft_file']}_{self.args['ref_var_ratio']}"
            )
        else:
            self.output_path = base_output_path
            
        check_directory(self.output_path)
        
        self.model_save_path = os.path.join(self.output_path, self.args["model_save_path"])
        self.hyperparams_outfile = os.path.join(self.output_path, self.args["hyperparams_log"])
        self.log_file = os.path.join(self.output_path, self.args["log_file_train"])
        self.checkpoint_path = os.path.join(self.output_path, "checkpoint.pth")
    
    def _setup_logging(self):
        """设置日志"""
        mode = "a" if self.args["checkpoint"] else "w"
        self.logger = setup_logger(self.log_file, mode=mode)
    
    def _validate_config(self):
        """验证配置"""
        required_keys = [
            "batch_size", "epochs", "lr", "weight_decay", "patience",
            "feature_size", "num_classes", "embed_dim", "patch_size",
            "window_size", "n_routed_experts", "n_activated_experts",
            "n_expert_groups", "n_limited_groups", "score_func",
            "route_scale", "moe_inter_dim", "n_shared_experts",
            "depths", "num_heads", "drop_rate", "drop_path_rate",
            "attn_drop_rate", "pct_start", "num_workers"
        ]
        
        for key in required_keys:
            if key not in self.args:
                raise ValueError(f"Missing required config key: {key}")
    
    def get_batch_size(self):
        """获取批量大小"""
        return self.args["ft_batch_size"] if self.args["ft"] else self.args["batch_size"]
    
    def get_epochs(self):
        """获取训练轮数"""
        return self.args["ft_epochs"] if self.args["ft"] else self.args["epochs"]
    
    def get_lr(self):
        """获取学习率"""
        return self.args["ft_lr"] if self.args["ft"] else self.args["lr"]
    
    def get_patience(self):
        """获取耐心值"""
        return self.args["ft_patience"] if self.args["ft"] else self.args["patience"]
    
    def log_hyperparams(self, model, optimizer, best_metrics):
        """记录超参数"""
        hyperparams = {
            "model": model.__class__.__name__,
            "optimizer": type(optimizer).__name__,
            "best_val_loss": best_metrics["best_val_loss"],
            "best_val_acc": best_metrics["best_val_acc"],
            "best_val_f1_score_all": best_metrics["best_val_f1_score"],
            "best_val_f1_score_snp": best_metrics["best_val_f1_score_snp"],
            "best_val_f1_score_indel": best_metrics["best_val_f1_score_indel"],
            "epochs": self.args["epochs"],
            "best_epoch": best_metrics["best_epoch"],
            "batch_size": self.args["batch_size"],
            "depths": self.args["depths"],
            "num_heads": self.args["num_heads"],
            "embed_dim": self.args["embed_dim"],
            "patch_size": self.args["patch_size"],
            "window_size": self.args["window_size"],
            "n_routed_experts": self.args["n_routed_experts"],
            "n_activated_experts": self.args["n_activated_experts"],
            "n_expert_groups": self.args["n_expert_groups"],
            "n_limited_groups": self.args["n_limited_groups"],
            "score_func": self.args["score_func"],
            "route_scale": self.args["route_scale"],
            "moe_inter_dim": self.args["moe_inter_dim"],
            "n_shared_experts": self.args["n_shared_experts"],
            "drop_rate": self.args["drop_rate"],
            "drop_path_rate": self.args["drop_path_rate"],
            "attn_drop_rate": self.args["attn_drop_rate"],
            "learning_rate": self.args["lr"],
            "weight_decay": self.args["weight_decay"],
            "pct_start": self.args["pct_start"],
        }
        
        hyperparams_df = pd.DataFrame([hyperparams])
        
        if os.path.exists(self.hyperparams_outfile):
            book = load_workbook(self.hyperparams_outfile)
            if "Transformer" in book.sheetnames:
                sheet = book["Transformer"]
                start_row = sheet.max_row
            else:
                start_row = 0
            with pd.ExcelWriter(
                self.hyperparams_outfile,
                mode="a",
                engine="openpyxl",
                if_sheet_exists="overlay",
            ) as writer:
                hyperparams_df.to_excel(
                    writer, 
                    sheet_name="Transformer", 
                    index=False, 
                    header=not ("Transformer" in load_workbook(self.hyperparams_outfile).sheetnames), 
                    startrow=start_row
                )
        else:
            with pd.ExcelWriter(
                self.hyperparams_outfile,
                mode="w",
                engine="openpyxl",
            ) as writer:
                hyperparams_df.to_excel(writer, sheet_name="Transformer", index=False, header=True)
    
    def save_training_process(self, train_process_data, epoch):
        """保存训练过程数据"""
        train_process = pd.DataFrame(
            data={
                "epoch": range(1, epoch + 2),
                "train_loss_list": train_process_data["train_loss_list"],
                "val_loss_list": train_process_data["val_loss_list"],
                "train_sensitivity_list": train_process_data["train_sensitivity_list"],
                "val_sensitivity_list": train_process_data["val_sensitivity_list"],
                "train_precision_list": train_process_data["train_precision_list"],
                "val_precision_list": train_process_data["val_precision_list"],
                "train_f1_score_list": train_process_data["train_f1_score_list"],
                "val_f1_score_list": train_process_data["val_f1_score_list"],
            },
        )
        train_process.to_excel(f"{self.output_path}/train_data.xlsx", index=False)
    
    def log_training_time(self, start_time):
        """记录训练时间"""
        consum_all = time.time() - start_time
        hours, remainder = divmod(consum_all, 3600)
        minutes, seconds = divmod(remainder, 60)
        self.logger.info(f"train time: {hours:.0f}h{minutes:.0f}m{seconds:.0f}s")